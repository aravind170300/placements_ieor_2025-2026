import streamlit as st
import pandas as pd
import openpyxl
import re

# 1. Setup the Webpage
st.set_page_config(page_title="Placements - IEOR 2025 - 2026", layout="wide")
st.title("📊 Placements - IEOR 2025-2026")
st.markdown("The companies listed here is open to IEOR department during the placement season 2025-2026 duing the Phase 1 alone. View JD for detailed description.")
st.markdown("Use the filters on the left to search and refine the data.")

# 2. Load the Data
@st.cache_data
def load_data():

    wb = openpyxl.load_workbook("placements_ieor_sheet.xlsx")
    ws = wb.active

    header_row = 3
    col_ctc = col_gross = None
    for cell in ws[header_row]:
        if cell.value == 'CTC(p.a)':
            col_ctc = cell.column
        if cell.value == 'Gross(p.a)':
            col_gross = cell.column

    def extract_symbol(fmt):
        match = re.search(r'\[\$(.+?)\]', fmt)
        return match.group(1) if match else '₹'

    currency_map = {}
    for row in ws.iter_rows(min_row=header_row + 1):
        r = row[0].row
        ctc_sym = extract_symbol(ws.cell(r, col_ctc).number_format) if col_ctc else '₹'
        gross_sym = extract_symbol(ws.cell(r, col_gross).number_format) if col_gross else '₹'
        currency_map[r] = (ctc_sym, gross_sym)

    df_raw = pd.read_excel("placements_ieor_sheet.xlsx", skiprows=2)
    df_raw = df_raw.loc[:, ~df_raw.columns.str.startswith('Unnamed')]

    was_nan = df_raw['CTC(p.a)'].isna().tolist()

    cols_to_fill = ['S.No', 'Company Name', 'Sector', 'Category', 'CTC(p.a)', 'Gross(p.a)']
    df_raw[cols_to_fill] = df_raw[cols_to_fill].ffill()

    df_raw['S.No'] = df_raw['S.No'].astype(int)

    ctc_syms = [currency_map.get(header_row + 1 + i, ('₹', '₹'))[0] for i in range(len(df_raw))]
    gross_syms = [currency_map.get(header_row + 1 + i, ('₹', '₹'))[1] for i in range(len(df_raw))]

    for i in range(1, len(ctc_syms)):
        if was_nan[i]:
            ctc_syms[i] = ctc_syms[i - 1]
            gross_syms[i] = gross_syms[i - 1]

    # Store raw numeric values for sorting before formatting
    df_raw['_ctc_raw'] = df_raw['CTC(p.a)'].copy()
    df_raw['_gross_raw'] = df_raw['Gross(p.a)'].copy()

    def format_currency(amount, symbol):
        if pd.isna(amount) or amount == '':
            return ''
        amount = int(float(amount))
        if symbol == '₹':
            s = str(amount)
            if len(s) <= 3:
                return f'₹{s}'
            last3 = s[-3:]
            rest = s[:-3]
            parts = []
            while len(rest) > 2:
                parts.append(rest[-2:])
                rest = rest[:-2]
            if rest:
                parts.append(rest)
            parts.reverse()
            return f'₹{",".join(parts)},{last3}'
        else:
            return f'{symbol}{amount:,}'

    df_raw['CTC(p.a)'] = [format_currency(v, s) for v, s in zip(df_raw['CTC(p.a)'], ctc_syms)]
    df_raw['Gross(p.a)'] = [format_currency(v, s) for v, s in zip(df_raw['Gross(p.a)'], gross_syms)]

    return df_raw

try:
    df = load_data()
except Exception as e:
    st.error(f"⚠️ Error loading file: {e}. Please check the filename and location.")
    st.stop()

# 3. Sidebar Filters
st.sidebar.header("🔍 Filter Options")

# --- Column visibility ---
all_columns = [c for c in df.columns if not c.startswith('_')]
selected_columns = st.sidebar.multiselect(
    "📋 Show/Hide Columns:",
    options=all_columns,
    default=all_columns
)
if not selected_columns:
    st.warning("Please select at least one column to display.")
    st.stop()

# --- Sort options ---
st.sidebar.markdown("---")
sort_col = st.sidebar.selectbox("↕️ Sort by:", options=["None"] + all_columns)
if sort_col != "None":
    sort_order = st.sidebar.radio("Order:", ["Ascending ↑", "Descending ↓"], horizontal=True)
    ascending = sort_order == "Ascending ↑"

# --- Master text search ---
st.sidebar.markdown("---")
search_query = st.sidebar.text_input("Search all text data:")

filtered_df = df.copy()

if search_query:
    mask = filtered_df.astype(str).apply(lambda x: x.str.contains(search_query, case=False, na=False))
    filtered_df = filtered_df[mask.any(axis=1)]

# --- Dropdown filters for categorical columns ---
categorical_cols = [col for col in df.columns if df[col].nunique() < 20 and df[col].dtype == 'object'
                    and col not in ('Job Description',) and not col.startswith('_')]

for col in categorical_cols:
    options = sorted(df[col].dropna().unique().tolist())
    selected = st.sidebar.multiselect(f"Filter by {col}:", options=options, default=options)
    filtered_df = filtered_df[filtered_df[col].isin(selected)]

# --- Apply sorting ---
if sort_col != "None":
    # Use raw numeric column for CTC/Gross sorting, display col otherwise
    sort_key = '_ctc_raw' if sort_col == 'CTC(p.a)' else '_gross_raw' if sort_col == 'Gross(p.a)' else sort_col
    filtered_df = filtered_df.sort_values(by=sort_key, ascending=ascending, na_position='last')

# 4. Make Job Description clickable links
def make_link(filename):
    if pd.isna(filename) or str(filename).strip() == '':
        return ''
    base_url = "https://raw.githubusercontent.com/aravind170300/placements_ieor_2025-2026/main/company/"
    return f'<a href="{base_url}{filename}" target="_blank">📄 View JD</a>'

display_df = filtered_df[selected_columns].copy()
if 'Job Description' in display_df.columns:
    display_df['Job Description'] = display_df['Job Description'].apply(make_link)

# 5. Display the Final Table
st.write(f"**Showing {len(filtered_df)} results | {len(selected_columns)} columns:**")
st.write(
    display_df.to_html(escape=False, index=False),
    unsafe_allow_html=True
)
